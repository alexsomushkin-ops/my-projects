from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import os
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import time
from typing import List, Dict


# Constants
SCROLL_PAUSE_TIME = 0.3
RETRIES = 5
DELAY = 2
IMAGE_DIR = 'images'


def get_user_input() -> str:
    """Get search phrase and optional city from user input."""
    searched_phrase = f'q-{input("Enter search phrase and press 'Enter': ")}'
    city_search = input("Specify city? Enter 'Y' or 'N': ").lower()
    searched_city = f"{input('Enter city name (optional): ')}" if city_search == 'y' else ""
    return f"https://www.olx.ua/uk/list/{searched_city}/{searched_phrase}/"


def slow_scroll_page(driver: webdriver, scroll_pause: float = SCROLL_PAUSE_TIME):
    """Scroll the webpage slowly, focusing on every fourth card briefly."""
    last_height = driver.execute_script("return document.body.scrollHeight")
    cards = driver.find_elements(By.CSS_SELECTOR, 'div[data-cy="l-card"]')

    for i, card in enumerate(cards):
        driver.execute_script("arguments[0].scrollIntoView();", card)
        if i % 5 == 0:
            time.sleep(0.2)
        time.sleep(scroll_pause)

    while True:
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(scroll_pause)
        new_height = driver.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height


def download_image(image_url: str, filename: str, retries: int = RETRIES, delay: int = DELAY) -> bool:
    """Download image with retries and handle failures gracefully."""
    for attempt in range(retries):
        try:
            response = requests.get(image_url, timeout=10)
            if response.status_code == 200:
                with open(filename, 'wb') as handler:
                    handler.write(response.content)
                return True
            print(f"Failed to download {image_url}. Status: {response.status_code}")
        except requests.RequestException as e:
            print(f"Error downloading {image_url}: {e}")
        print(f"Retry {attempt + 1}/{retries} for {image_url}")
        time.sleep(delay)
    return False


def scrape_olx(driver: webdriver, websitelink: str) -> List[Dict[str, str]]:
    """Scrape data from OLX site."""
    driver.get(websitelink)
    WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div[data-cy="l-card"]')))
    slow_scroll_page(driver)

    # Determine number of pages
    pages = driver.find_elements(By.CSS_SELECTOR, 'li[data-testid="pagination-list-item"]')
    number_of_pages = int(pages[-1].text) if pages else 1

    data = []

    # Ensure image directory exists
    os.makedirs(IMAGE_DIR, exist_ok=True)

    for page in range(1, number_of_pages + 1):
        cards = driver.find_elements(By.CSS_SELECTOR, 'div[data-cy="l-card"]')
        for card in cards:
            try:
                name = card.find_element(By.TAG_NAME, "h6").text
                price = card.find_element(By.CSS_SELECTOR, 'p[data-testid="ad-price"]').text
                location_time = card.find_element(By.CSS_SELECTOR, 'p[data-testid="location-date"]').text.split(" - ")
                location, timeposted = location_time if len(location_time) == 2 else ("", "")
                image_url = card.find_element(By.CSS_SELECTOR, 'img[src]').get_attribute("src")
                ad_url = card.find_element(By.CSS_SELECTOR, 'a[href]').get_attribute("href")

                img_filename = os.path.join(IMAGE_DIR, f"{name.replace(' ', '_')}.jpg")
                if download_image(image_url, img_filename):
                    data.append({
                        "Name": name,
                        "Price": price,
                        "Location": location,
                        "Time Posted": timeposted,
                        "Image File": img_filename,
                        "Ad URL": ad_url
                    })
            except Exception as e:
                print(f"Error processing card: {e}")

        if page < number_of_pages:
            driver.get(f'{websitelink}&page={page + 1}')
            WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div[data-cy="l-card"]')))
            slow_scroll_page(driver)

    return data


def save_to_excel(data: List[Dict[str, str]], filename: str = "ad_data_with_images.xlsx"):
    """Save scraped data to an Excel file with images."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ad Data"
    headers = ["Name", "Price", "Location", "Time Posted", "Image", "Ad URL"]
    ws.append(headers)

    for entry in data:
        ws.append([entry["Name"], entry["Price"], entry["Location"], entry["Time Posted"], "", entry["Ad URL"]])

        img = Image(entry["Image File"])
        img.width, img.height = 215, 152
        ws.add_image(img, f"E{ws.max_row}")
        ws.row_dimensions[ws.max_row].height = img.height

    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    wb.save(filename)
    print(f"Data saved to {filename}")


def main():
    websitelink = get_user_input()
    driver = webdriver.Edge()
    try:
        data = scrape_olx(driver, websitelink)
        save_to_excel(data)
    finally:
        driver.quit()


if __name__ == "__main__":
    main()
